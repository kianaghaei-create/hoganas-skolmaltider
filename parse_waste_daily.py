"""
Label-baserad daglig svinnparsare — identifierar rader via textetiketter
i stället för hårdkodade radnummer.

Stöder minst tre format:
  Format A (Förskola):
    Köks- & serveringssvinn (kg):  ← kombinerat på en rad
    Tallrikssvinn (kg):
    Totalt uppmätt matsvinn (kg):
    Serveringssvinn (%)
    Tallrikssvinn (%)
    Totalt uppmätt matsvinn (%)

  Format B (Skola/ÄO):
    Kökssvinn, ej uppdelat (kg):
    Serveringssvinn (kg):
    Tallrikssvinn (kg):
    Totalt uppmätt matsvinn (kg):
    Kökssvinn (%)
    Serveringssvinn (%)
    Tallrikssvinn (%)
    Totalt uppmätt matsvinn (%)
"""
from pathlib import Path
import re
import openpyxl
import pandas as pd
from datetime import datetime, timedelta

WASTE_DIR = Path("Data/Matsvinn 2025")
OUT       = Path("Data/processed")
OUT.mkdir(parents=True, exist_ok=True)

WEEKDAYS = ['Måndag', 'Tisdag', 'Onsdag', 'Torsdag', 'Fredag']


# ── Hjälpfunktioner ───────────────────────────────────────────────────────────

def to_float(v):
    if v is None:
        return None
    s = str(v).replace(',', '.').replace('%', '').strip()
    if s.lower() in ('none', 'värde saknas', 'automatisk beräkning', ''):
        return None
    try:
        f = float(s)
        return f if f == f else None  # NaN guard
    except (TypeError, ValueError):
        return None


def normalize_label(v) -> str:
    """Gör etiketten jämförbar: gemener, trim, ta bort kolon i slutet."""
    if v is None:
        return ''
    s = str(v).lower().strip()
    # Komprimera interna whitespace-sekvenser
    s = re.sub(r'\s+', ' ', s)
    # Ta bort avslutande kolon
    s = s.rstrip(':').strip()
    return s


def week_to_monday(year, week):
    try:
        return datetime.strptime(f'{int(year)}-W{int(week):02d}-1', '%G-W%V-%u').date()
    except Exception:
        return None


def parse_week_number(val):
    if val is None:
        return None
    m = re.search(r'(\d+)', str(val))
    return int(m.group(1)) if m else None


# ── Huvud-parsningsfunktion ───────────────────────────────────────────────────

def parse_file(fpath: Path) -> list[dict]:
    enhet_default = fpath.stem
    records = []

    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  SKIP {enhet_default}: {e}")
        return records

    for shname in wb.sheetnames:
        if shname in ('Sammanställning', 'ESRI_MAPINFO_SHEET'):
            continue

        ws   = wb[shname]
        rows = list(ws.iter_rows(max_row=30, max_col=10, values_only=True))
        if len(rows) < 10:
            continue

        # ── Meta-rader (fasta positioner — metadata, inte mätvärden) ──────────
        enhet    = str(rows[0][1]).strip() if rows[0][1] else enhet_default
        if not enhet or enhet.lower() == 'none':
            enhet = enhet_default
        week_num = parse_week_number(rows[1][1])
        year_raw = rows[1][3]
        year     = int(year_raw) if year_raw and str(year_raw).isdigit() else 2025
        if week_num is None:
            continue
        monday       = week_to_monday(year, week_num)
        portionsvikt = to_float(rows[2][1])

        # ── Hitta dagkolumner via header-rad ──────────────────────────────────
        day_col: dict[str, int] = {}
        for row in rows[:8]:
            for j, cell in enumerate(row):
                norm = normalize_label(cell)
                if norm in ('måndag', 'tisdag', 'onsdag', 'torsdag', 'fredag'):
                    day_col[norm.capitalize()] = j
        if not day_col:
            continue

        # ── Bygg etikett → rad-index-mappning ─────────────────────────────────
        label_to_row: dict[str, list] = {}
        for i, row in enumerate(rows):
            lbl = normalize_label(row[0])
            if lbl:
                # Spara alla rader med denna normaliserade etikett (kan vara duplikat)
                label_to_row.setdefault(lbl, []).append((i, row))

        def get_vals(keys: list[str]) -> list | None:
            """Returnera första matchande rad, söker i ordning."""
            for key in keys:
                key_n = normalize_label(key)
                # Exakt matchning
                if key_n in label_to_row:
                    return label_to_row[key_n][0][1]
                # Partiell matchning (alla ord i nyckeln måste finnas)
                for lbl, row_list in label_to_row.items():
                    if all(w in lbl for w in key_n.split()):
                        return row_list[0][1]
            return None

        def day_val(row, weekday: str):
            if row is None:
                return None
            col = day_col.get(weekday)
            if col is None:
                return None
            try:
                return to_float(row[col])
            except IndexError:
                return None

        # ── Detektera format ──────────────────────────────────────────────────
        # Förskola: har "köks- & serveringssvinn" (kombinerad), ej separat köks+serv
        has_combined = any(
            'köks' in lbl and 'serveringssvinn' in lbl and '%)' not in lbl
            for lbl in label_to_row
        )
        has_separate_koks = any(
            'kökssvinn' in lbl and '%)' not in lbl and 'server' not in lbl
            for lbl in label_to_row
        )

        # ── Hämta datarader via etiketter ─────────────────────────────────────
        r_matratt    = get_vals(['maträtt'])
        r_kommentar  = get_vals(['kommentar'])

        # Beställda: välj raden utan "(kg)" — den med portionsantal
        r_bestallda  = None
        for lbl, row_list in label_to_row.items():
            if 'beställda portioner' in lbl and '(kg)' not in lbl:
                r_bestallda = row_list[0][1]
                break

        r_serverade  = get_vals([
            'antal serverade portioner',
            'serverade portioner',
        ])

        # Kg-rader
        if has_combined:
            r_kok_serv   = get_vals(['köks- & serveringssvinn (kg)', 'köks och serveringssvinn'])
            r_koks_kg    = None
            r_serv_kg    = None
        else:
            r_kok_serv   = None
            r_koks_kg    = get_vals(['kökssvinn, ej uppdelat (kg)', 'kökssvinn (kg)'])
            r_serv_kg    = get_vals(['serveringssvinn (kg)'])

        r_tallrik_kg = get_vals(['tallrikssvinn (kg)'])
        r_total_kg   = get_vals([
            'totalt uppmätt matsvinn (kg)',
            'totalt svinn (kg)',
        ])

        # Procent-rader
        r_koks_pct    = get_vals(['kökssvinn (%)'])
        r_serv_pct    = get_vals(['serveringssvinn (%)'])
        r_tallrik_pct = get_vals(['tallrikssvinn (%)'])
        r_total_pct   = get_vals([
            'totalt uppmätt matsvinn (%)',
            'totalt (%)',
        ])

        def day_str(row, weekday: str) -> str:
            """Rå sträng (ej to_float) för textrader som maträtt/kommentar."""
            if row is None:
                return ''
            col = day_col.get(weekday)
            if col is None:
                return ''
            try:
                v = row[col]
                return str(v).strip() if v is not None and str(v).lower() not in ('none', '') else ''
            except IndexError:
                return ''

        # ── Bygg en rad per dag ───────────────────────────────────────────────
        for i, weekday in enumerate(WEEKDAYS):
            matratt   = day_str(r_matratt,   weekday)
            kommentar = day_str(r_kommentar, weekday)
            bestallda = day_val(r_bestallda, weekday)
            serverade = day_val(r_serverade, weekday)

            koks_serv_combined = day_val(r_kok_serv,    weekday)
            koks_kg            = day_val(r_koks_kg,     weekday)
            serv_kg            = day_val(r_serv_kg,     weekday)
            tallrik_kg         = day_val(r_tallrik_kg,  weekday)
            total_kg           = day_val(r_total_kg,    weekday)

            koks_pct    = day_val(r_koks_pct,    weekday)
            serv_pct    = day_val(r_serv_pct,    weekday)
            tallrik_pct = day_val(r_tallrik_pct, weekday)
            total_pct   = day_val(r_total_pct,   weekday)

            # Skippa dagar utan någon meningsfull data
            if (
                not matratt
                and bestallda is None
                and (koks_kg or 0) == 0
                and (koks_serv_combined or 0) == 0
                and (total_kg or 0) == 0
            ):
                continue

            datum = (monday + timedelta(days=i)) if monday else None

            records.append({
                'enhet':                       enhet,
                'fil':                         fpath.name,
                'blad':                        shname,
                'vecka':                       week_num,
                'ar':                          year,
                'datum':                       str(datum) if datum else None,
                'veckodag':                    weekday,
                'matratt':                     matratt or None,
                'kommentar':                   kommentar or None,
                'portionsvikt_g':              portionsvikt,
                'bestallda_portioner':         bestallda,
                'serverade_portioner':         serverade,
                # Kg-svinn — format A (förskola) använder kombinerat fält
                'kok_och_serveringssvinn_kg':  koks_serv_combined,  # Format A
                'kokssvinn_kg':                koks_kg,              # Format B
                'serveringssvinn_kg':          serv_kg,              # Format B
                'tallrikssvinn_kg':            tallrik_kg,
                'totalt_svinn_kg':             total_kg,
                # Procent-svinn
                'kokssvinn_pct':               koks_pct,
                'serveringssvinn_pct':         serv_pct,
                'tallrikssvinn_pct':           tallrik_pct,
                'totalt_svinn_pct':            total_pct,
                # Format-flagga
                'format':                      'förskola' if has_combined else 'skola_ao',
            })

    wb.close()
    return records


# ── Kör på alla svinnfiler ────────────────────────────────────────────────────
all_records = []
files = sorted(WASTE_DIR.glob('*.xlsx'))
print(f"Processar {len(files)} svinnfiler (label-baserad parser)...\n")

for fpath in files:
    recs = parse_file(fpath)
    fmt  = recs[0]['format'] if recs else '?'
    print(f"  {fpath.stem}: {len(recs)} dagsrader  [{fmt}]")
    all_records.extend(recs)

df = pd.DataFrame(all_records)

# ── Sanity-filter: ta bort uppenbart felaktiga värden ────────────────────────
# Max 500 kg/dag per enhet är rimligt
df = df[df['totalt_svinn_kg'].fillna(0) < 500]

# ── Normalisera rättnamn ──────────────────────────────────────────────────────
df['matratt_norm'] = df['matratt'].str.lower().str.strip()
df.loc[df['matratt_norm'] == 'none', 'matratt_norm'] = None
df.loc[df['matratt_norm'] == '',     'matratt_norm'] = None

# Normalisera enhetnamn (ta bort extra whitespace)
df['unit_name'] = df['enhet'].str.strip()

# ── Spara ─────────────────────────────────────────────────────────────────────
out_path = OUT / 'food_waste_daily_v2.csv'
df.to_csv(out_path, index=False, encoding='utf-8-sig')

# ── Rapport ───────────────────────────────────────────────────────────────────
print(f"\n✅ {len(df)} rader sparade → {out_path}")
print(f"   Enheter:               {df['unit_name'].nunique()}")
print(f"   Datum-range:           {df['datum'].min()} → {df['datum'].max()}")
print(f"   Format A (förskola):   {(df['format']=='förskola').sum()} rader")
print(f"   Format B (skola/ÄO):   {(df['format']=='skola_ao').sum()} rader")
print(f"   Med rättnamn:          {df['matratt_norm'].notna().sum()} rader")
print(f"   Med totalt_svinn_kg:   {df['totalt_svinn_kg'].notna().sum()} rader")
print(f"   Med totalt_svinn_pct:  {df['totalt_svinn_pct'].notna().sum()} rader")

print("\n--- totalt_svinn_kg per format ---")
print(df.groupby('format')['totalt_svinn_kg'].agg(['sum','count','mean']).round(2))

print("\n--- totalt_svinn_pct NaN per format ---")
print(df.groupby('format')['totalt_svinn_pct'].apply(lambda x: x.isna().sum()).rename('nan_pct'))

print("\n--- Rader per enhet ---")
summary = df.groupby('unit_name').agg(
    dagar         =('datum', 'count'),
    format        =('format', 'first'),
    svinn_kg_sum  =('totalt_svinn_kg', 'sum'),
    pct_nan_count =('totalt_svinn_pct', lambda x: x.isna().sum()),
).round(1).sort_values('dagar', ascending=False)
print(summary.to_string())
