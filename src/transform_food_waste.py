from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd


def _week_num(text: str) -> int | None:
    m = re.search(r"vecka\s*(\d{1,2})", str(text).lower())
    return int(m.group(1)) if m else None


def _unit_from_file(file_path: Path) -> str:
    return file_path.stem  # e.g. "Kullagymnasiet"


def parse_food_waste_file(file_path: Path) -> pd.DataFrame:
    """
    Read the 'Sammanställning' sheet from a food waste workbook.
    Columns (raw): Kökets namn | Start-vecka | Kökssvinn% | Serveringssvinn% |
                   Tallrikssvinn% | Totalt% | Totalt kg | Beställda | Serverade | Överbeställning
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return pd.DataFrame()

    # Find the Sammanställning sheet (case-insensitive)
    sam_sheet = None
    for name in wb.sheetnames:
        if "sammanst" in name.lower():
            sam_sheet = name
            break
    if not sam_sheet:
        return pd.DataFrame()

    ws = wb[sam_sheet]
    rows_data = list(ws.iter_rows(values_only=True))

    if not rows_data:
        return pd.DataFrame()

    # Find header row (contains "vecka" or "kökets namn")
    header_idx = None
    for i, row in enumerate(rows_data[:10]):
        row_str = " ".join(str(v).lower() for v in row if v is not None)
        if "vecka" in row_str or "kokets" in row_str or "kök" in row_str.lower():
            header_idx = i
            break

    if header_idx is None:
        return pd.DataFrame()

    def to_float(v):
        try:
            f = float(v)
            return f if f == f else None  # NaN guard
        except (TypeError, ValueError):
            return None

    # Map header → column index dynamically
    header_row = [str(v).lower().strip().replace("\n", " ") if v else "" for v in rows_data[header_idx]]

    def col(keywords):
        """Return index of first header cell containing ALL keywords."""
        for i, h in enumerate(header_row):
            if all(kw in h for kw in keywords):
                return i
        return None

    idx_unit     = col(["kökets"]) or 0
    idx_week     = col(["vecka"]) or 1
    idx_kitchen  = col(["kökssvinn", "%"]) if col(["kökssvinn", "%"]) is not None else None
    idx_serving  = col(["serveringssvinn"]) if col(["serveringssvinn"]) is not None else None
    idx_combined = col(["köks", "server"]) if (idx_kitchen is None) else None  # merged column
    idx_plate    = col(["tallrikssvinn"]) or None
    idx_total_pct= col(["totalt", "%"]) or col(["totalt uppmätt matsvinn (%)"]) or None
    idx_total_kg = col(["totalt", "kg"]) or col(["totalt uppmätt matsvinn (kg)"]) or None
    idx_ordered  = col(["beställda"]) or None
    idx_served   = col(["serverade"]) or None
    idx_over     = col(["överbeställning"]) or None

    records = []
    for row in rows_data[header_idx + 1:]:
        if not any(v is not None for v in row):
            continue
        vals = list(row)

        def get(idx):
            if idx is None:
                return None
            try:
                return vals[idx]
            except IndexError:
                return None

        unit_name = str(get(idx_unit)).strip() if get(idx_unit) else _unit_from_file(file_path)
        if not unit_name or unit_name.lower() in ("none", "kökets namn", ""):
            continue

        week = _week_num(str(get(idx_week))) if get(idx_week) else None

        kitchen_waste_pct  = to_float(get(idx_kitchen))
        serving_waste_pct  = to_float(get(idx_serving))
        combined_waste_pct = to_float(get(idx_combined))  # förskola: köks+serving merged
        plate_waste_pct    = to_float(get(idx_plate))
        total_waste_pct    = to_float(get(idx_total_pct))
        total_waste_kg     = to_float(get(idx_total_kg))
        ordered_portions   = to_float(get(idx_ordered))
        served_portions    = to_float(get(idx_served))
        over_order_ratio   = to_float(get(idx_over))

        # Skip empty weeks
        if (total_waste_kg or 0) == 0 and (ordered_portions or 0) == 0 and (served_portions or 0) == 0:
            continue

        records.append({
            "unit_name":            unit_name,
            "year":                 2025,
            "week":                 week,
            "kitchen_waste_pct":    kitchen_waste_pct,
            "serving_waste_pct":    serving_waste_pct,
            "combined_waste_pct":   combined_waste_pct,
            "plate_waste_pct":      plate_waste_pct,
            "total_waste_pct":      total_waste_pct,
            "total_waste_kg":       total_waste_kg,
            "ordered_portions":     ordered_portions,
            "served_portions":      served_portions,
            "over_order_ratio":     over_order_ratio,
            "source_file":          str(file_path),
            "source_sheet":         sam_sheet,
            "detected_category":    "food_waste",
        })

    return pd.DataFrame(records) if records else pd.DataFrame()


def transform_food_waste(df: pd.DataFrame, file_path: Path, sheet_name: str) -> pd.DataFrame:
    """
    Called per-sheet by pipeline — skip here, real work done in parse_food_waste_file.
    """
    return pd.DataFrame()
