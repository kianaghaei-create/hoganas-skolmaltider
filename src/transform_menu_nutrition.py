from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd

WEEKDAY_MAP = {
    "måndag": 1, "tisdag": 2, "onsdag": 3, "torsdag": 4, "fredag": 5,
    "lördag": 6, "söndag": 7,
}

MONTH_MAP = {
    "januari": 1, "februari": 2, "mars": 3, "april": 4,
    "maj": 5, "juni": 6, "juli": 7, "augusti": 8,
    "september": 9, "oktober": 10, "november": 11, "december": 12,
}


def _norm(s: str) -> str:
    import unicodedata
    s = str(s).lower().strip()
    s = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in s if not unicodedata.combining(ch))


def _parse_week_header(text: str) -> tuple[int | None, int | None]:
    """'MATSEDEL JANUARI VECKA 2' → (month=1, week=2)"""
    t = _norm(text)
    week = None
    month = None
    m = re.search(r"vecka\s*(\d+)", t)
    if m:
        week = int(m.group(1))
    for swe, num in MONTH_MAP.items():
        if swe in t:
            month = num
            break
    return month, week


def _parse_weekday_date(text: str) -> tuple[int | None, str | None]:
    """'Måndag 6/1' → (weekday=1, date_str='6/1')"""
    t = _norm(text)
    for day_swe, day_num in WEEKDAY_MAP.items():
        if t.startswith(day_swe):
            date_part = re.search(r"(\d{1,2}/\d{1,2})", text)
            return day_num, date_part.group(1) if date_part else None
    return None, None


def _parse_dish(text: str) -> tuple[str | None, str | None]:
    """
    'Dagens lunch: Spaghetti med köttfärssås' → ('dagens_lunch', 'Spaghetti med köttfärssås')
    'A-kost alt 1: Lokal wienerkorv...'       → ('a_kost_alt_1', 'Lokal wienerkorv...')
    """
    if ":" not in text:
        return None, None
    parts = text.split(":", 1)
    dish_type = re.sub(r"[^a-z0-9]+", "_", _norm(parts[0])).strip("_")
    dish_name = parts[1].strip()
    if not dish_name:
        return None, None
    return dish_type, dish_name


def parse_menu_file(file_path: Path, menu_type: str) -> pd.DataFrame:
    """
    Parse an entire menu workbook (one sheet per week).
    Returns long-format: week | month | year | weekday | date_str | dish_type | dish_name | menu_type
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return pd.DataFrame()

    rows = []
    year = 2025  # All data is 2025

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Data may be in col A or col B depending on formatting — collect all non-empty cells
        values = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None and str(cell).strip() and str(cell).strip() != "":
                    values.append(str(cell).strip())
                    break  # take first non-empty cell per row

        month, week = None, None
        current_weekday = None
        current_date_str = None

        for cell_val in values:
            text = str(cell_val).strip()
            t = _norm(text)

            # Week header
            if "matsedel" in t and "vecka" in t:
                month, week = _parse_week_header(text)
                continue

            # Weekday line
            weekday, date_str = _parse_weekday_date(text)
            if weekday is not None:
                current_weekday = weekday
                current_date_str = date_str
                continue

            # Dish line
            if ":" in text and week is not None:
                dish_type, dish_name = _parse_dish(text)
                if dish_type and dish_name:
                    # Skip placeholder entries
                    if dish_name.lower() in ("skollov", "stängt", "special alt 1", "special alt 2", ""):
                        continue
                    rows.append({
                        "year": year,
                        "month": month,
                        "week": week,
                        "weekday": current_weekday,
                        "date_str": current_date_str,
                        "dish_type": dish_type,
                        "dish_name": dish_name,
                        "menu_type": menu_type,
                        "source_file": str(file_path),
                        "source_sheet": sheet_name,
                        "detected_category": "menu_nutrition",
                    })

    return pd.DataFrame(rows) if rows else pd.DataFrame()


def transform_menu_nutrition(df: pd.DataFrame, file_path: Path, sheet_name: str) -> pd.DataFrame:
    """
    Called per sheet by the pipeline — but menu files need full-workbook parsing.
    We return empty here and let the pipeline handle it via parse_menu_file directly.
    This function is kept for interface compatibility but real work is in parse_menu_file.
    """
    return pd.DataFrame()
